import io
import re
from datetime import datetime
from typing import Any

from models.report import ReportRequest


def generate(request: ReportRequest):
    """根据请求生成报表文件"""
    fmt = request.format.lower()
    if fmt not in ("excel", "pdf"):
        raise ValueError(f"不支持的报表格式: {request.format}")

    report_type = request.report_type
    data = request.data or []
    params = request.params or {}

    generator = REPORT_REGISTRY.get(report_type)
    if generator is None:
        # 未知类型使用通用报表
        generator = generate_generic

    headers, rows, title = generator(report_type, data, params)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    if fmt == "excel":
        content = _generate_excel(title, headers, rows)
        filename = f"{report_type}_{timestamp}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _generate_pdf(title, headers, rows)
        filename = f"{report_type}_{timestamp}.pdf"
        content_type = "application/pdf"

    return content, filename, content_type


# ===================== 各报表类型实现 =====================

def generate_inventory(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """库存报表"""
    title = "库存报表"
    headers = [
        ("warehouseName", "仓库"),
        ("goodsCode", "商品编码"),
        ("skuCode", "SKU编码"),
        ("goodsName", "商品名称"),
        ("goodsSpec", "规格"),
        ("unit", "单位"),
        ("quantity", "库存数量"),
        ("availableQuantity", "可用数量"),
        ("lockedQuantity", "锁定数量"),
        ("avgCostPrice", "平均成本价"),
        ("totalAmount", "库存金额"),
    ]
    rows = [_map_row(d, headers) for d in data]
    return headers, rows, title


def generate_purchase_order(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """采购订单报表"""
    title = "采购订单报表"
    headers = [
        ("orderNo", "采购单号"),
        ("docName", "单据名称"),
        ("supplierName", "供应商"),
        ("orderDate", "订单日期"),
        ("deliveryDate", "预计交货日期"),
        ("totalAmount", "订单金额"),
        ("taxAmount", "税额"),
        ("discountAmount", "折扣金额"),
        ("payAmount", "实付金额"),
        ("orderStatus", "订单状态"),
        ("remark", "备注"),
    ]
    rows = []
    for d in data:
        row = _map_row(d, headers)
        row[9] = _map_purchase_status(d.get("orderStatus"))
        rows.append(row)
    return headers, rows, title


def generate_sale_order(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """销售订单报表"""
    title = "销售订单报表"
    headers = [
        ("orderNo", "销售单号"),
        ("docName", "单据名称"),
        ("customerName", "客户"),
        ("orderDate", "订单日期"),
        ("deliveryDate", "交货日期"),
        ("saleType", "销售类型"),
        ("totalAmount", "订单金额"),
        ("taxAmount", "税额"),
        ("discountAmount", "折扣金额"),
        ("payAmount", "实收金额"),
        ("orderStatus", "订单状态"),
        ("warehouseName", "仓库"),
    ]
    rows = []
    for d in data:
        row = _map_row(d, headers)
        row[5] = _map_sale_type(d.get("saleType"))
        row[10] = _map_sale_status(d.get("orderStatus"))
        rows.append(row)
    return headers, rows, title


def generate_stock_flow(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """库存流水报表"""
    title = "库存流水报表"
    headers = [
        ("warehouseName", "仓库"),
        ("goodsCode", "商品编码"),
        ("skuCode", "SKU编码"),
        ("goodsName", "商品名称"),
        ("sourceType", "业务类型"),
        ("sourceNo", "来源单号"),
        ("inQuantity", "入库数量"),
        ("outQuantity", "出库数量"),
        ("beforeQuantity", "变动前数量"),
        ("afterQuantity", "变动后数量"),
        ("costPrice", "成本价"),
        ("amount", "金额"),
        ("remark", "备注"),
    ]
    rows = [_map_row(d, headers) for d in data]
    return headers, rows, title


def generate_psi_summary(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """进销存汇总报表"""
    title = "进销存汇总报表"
    headers = [
        ("goodsCode", "商品编码"),
        ("skuCode", "SKU编码"),
        ("goodsName", "商品名称"),
        ("warehouseName", "仓库"),
        ("unit", "单位"),
        ("beginningQuantity", "期初库存"),
        ("inQuantity", "本期入库"),
        ("outQuantity", "本期出库"),
        ("currentQuantity", "期末库存"),
        ("availableQuantity", "可用库存"),
        ("purchaseAmount", "采购金额"),
        ("saleAmount", "销售金额"),
        ("stockAmount", "库存金额"),
    ]
    rows = [_map_row(d, headers) for d in data]
    return headers, rows, title


def generate_generic(report_type: str, data: list[dict[str, Any]], params: dict[str, Any]):
    """通用报表：按数据中的 key 生成"""
    title = f"{report_type} 报表"
    if not data:
        headers = [("message", "提示")]
        rows = [["暂无数据"]]
        return headers, rows, title

    keys = list(data[0].keys())
    headers = [(k, k) for k in keys]
    rows = [[_fmt_value(d.get(k)) for k in keys] for d in data]
    return headers, rows, title


REPORT_REGISTRY = {
    "inventory": generate_inventory,
    "purchase_order": generate_purchase_order,
    "sale_order": generate_sale_order,
    "stock_flow": generate_stock_flow,
    "psi_summary": generate_psi_summary,
}


def _map_row(data: dict[str, Any], headers: list[tuple[str, str]]) -> list[Any]:
    """按表头顺序从数据中提取字段"""
    return [_fmt_value(data.get(key)) for key, _ in headers]


def _fmt_value(value: Any) -> Any:
    """格式化单元格值"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value)
    # 日期时间字符串截断为日期
    if re.match(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}", s):
        return s[:10]
    return s


def _map_purchase_status(value: Any) -> str:
    """采购订单状态映射"""
    mapping = {
        1: "待审核",
        2: "已审核",
        3: "已取消",
        4: "已完成",
        5: "已入库",
        "PENDING": "待审核",
        "APPROVED": "已审核",
        "IN_STOCK": "已入库",
        "CANCELLED": "已取消",
        "COMPLETED": "已完成",
    }
    return mapping.get(value, str(value) if value is not None else "")


def _map_sale_status(value: Any) -> str:
    """销售订单状态映射"""
    mapping = {
        1: "待审核",
        2: "已审核",
        3: "已取消",
        4: "已完成",
        5: "已出库",
        "PENDING": "待审核",
        "APPROVED": "已审核",
        "OUT_STOCK": "已出库",
        "CANCELLED": "已取消",
        "COMPLETED": "已完成",
    }
    return mapping.get(value, str(value) if value is not None else "")


def _map_sale_type(value: Any) -> str:
    """销售类型映射"""
    mapping = {
        1: "零售",
        2: "批发",
        3: "团购",
        4: "电商",
        5: "其他",
        "RETAIL": "零售",
        "WHOLESALE": "批发",
        "GROUP": "团购",
        "ONLINE": "电商",
        "OTHER": "其他",
    }
    return mapping.get(value, str(value) if value is not None else "")


# ===================== Excel / PDF 生成 =====================

def _generate_excel(title: str, headers: list[tuple[str, str]], rows: list[list[Any]]) -> bytes:
    """生成 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 表头
    header_labels = [label for _, label in headers]
    ws.append(header_labels)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 数据行
    for row_data in rows:
        ws.append(row_data)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if col_idx <= 3 else "right")

    # 自动列宽（跳过合并单元格）
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = ws.cell(row=2, column=col_idx).column_letter
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _generate_pdf(title: str, headers: list[tuple[str, str]], rows: list[list[Any]]) -> bytes:
    """生成 PDF 文件"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 12))

    header_labels = [label for _, label in headers]
    table_data = [header_labels]
    for row in rows:
        table_data.append([str(v) for v in row])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()
